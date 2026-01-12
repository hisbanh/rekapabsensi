"""
Report Service Layer
Handles all business logic related to reporting and analytics
"""
from typing import List, Dict, Optional
from datetime import date, datetime, timedelta
from django.db.models import Q, Count, Avg
from django.core.paginator import Paginator
import csv
from io import StringIO, BytesIO

from ..models import (
    Student, AttendanceRecord, AttendanceStatus, Classroom,
    DailyAttendance, DaySchedule
)
from ..exceptions import ReportServiceError
from .schedule_service import ScheduleService
from .holiday_service import HolidayService


class ReportService:
    """Service class for report generation and analytics"""
    
    # ============================================
    # JP-Based Report Methods (New)
    # ============================================
    
    @staticmethod
    def generate_class_report(
        classroom: Classroom,
        start_date: date,
        end_date: date
    ) -> Dict:
        """
        Generate comprehensive attendance report for a classroom.
        
        Calculates totals and percentages for each student across all JP slots
        in the date range.
        
        Args:
            classroom: The classroom to generate report for
            start_date: Start of date range
            end_date: End of date range
            
        Returns:
            Dict containing:
            - classroom: Classroom object
            - period: Dict with start_date, end_date
            - students: List of student report data with totals and percentages
            - class_summary: Aggregate totals for the entire class
            - dates: List of dates in the range (school days only)
            
        Requirements: 5.3, 5.4
        """
        # Get all active students in the classroom
        students = Student.objects.filter(
            classroom=classroom,
            is_active=True
        ).order_by('name')
        
        # Get all school days in the date range (excluding holidays)
        school_dates = ReportService._get_school_dates(
            classroom, start_date, end_date
        )
        
        # Get all attendance records for this classroom in the date range
        attendances = DailyAttendance.objects.filter(
            student__classroom=classroom,
            date__range=[start_date, end_date]
        ).select_related('student')
        
        # Build attendance lookup: {student_id: {date: jp_statuses}}
        attendance_lookup = {}
        for attendance in attendances:
            student_id = str(attendance.student.id)
            if student_id not in attendance_lookup:
                attendance_lookup[student_id] = {}
            attendance_lookup[student_id][attendance.date] = attendance.jp_statuses
        
        # Calculate totals for each student
        student_reports = []
        class_total_h = 0
        class_total_s = 0
        class_total_i = 0
        class_total_a = 0
        class_total_jp = 0
        
        for student in students:
            student_id = str(student.id)
            student_attendance = attendance_lookup.get(student_id, {})
            
            # Calculate totals for this student
            total_h = 0
            total_s = 0
            total_i = 0
            total_a = 0
            total_jp = 0
            
            # Daily breakdown for this student
            daily_data = []
            
            for school_date in school_dates:
                jp_statuses = student_attendance.get(school_date, {})
                jp_count = ScheduleService.get_jp_count_for_date(school_date)
                
                day_h = 0
                day_s = 0
                day_i = 0
                day_a = 0
                
                for jp_num in range(1, jp_count + 1):
                    status = jp_statuses.get(str(jp_num), None)
                    if status == 'H':
                        day_h += 1
                        total_h += 1
                    elif status == 'S':
                        day_s += 1
                        total_s += 1
                    elif status == 'I':
                        day_i += 1
                        total_i += 1
                    elif status == 'A':
                        day_a += 1
                        total_a += 1
                    # If no status recorded, count as missing (not included in totals)
                    if status:
                        total_jp += 1
                
                daily_data.append({
                    'date': school_date,
                    'jp_count': jp_count,
                    'hadir': day_h,
                    'sakit': day_s,
                    'izin': day_i,
                    'alpa': day_a,
                    'has_record': bool(jp_statuses),
                    'summary': f"H:{day_h} S:{day_s} I:{day_i} A:{day_a}" if jp_statuses else "-"
                })
            
            # Calculate percentage
            attendance_percentage = round(
                (total_h / total_jp * 100), 2
            ) if total_jp > 0 else 0.0
            
            student_reports.append({
                'student': student,
                'total_hadir': total_h,
                'total_sakit': total_s,
                'total_izin': total_i,
                'total_alpa': total_a,
                'total_jp': total_jp,
                'attendance_percentage': attendance_percentage,
                'daily_data': daily_data,
            })
            
            # Add to class totals
            class_total_h += total_h
            class_total_s += total_s
            class_total_i += total_i
            class_total_a += total_a
            class_total_jp += total_jp
        
        # Calculate class average percentage
        class_attendance_percentage = round(
            (class_total_h / class_total_jp * 100), 2
        ) if class_total_jp > 0 else 0.0
        
        return {
            'classroom': classroom,
            'period': {
                'start_date': start_date,
                'end_date': end_date,
            },
            'students': student_reports,
            'class_summary': {
                'total_students': len(students),
                'total_hadir': class_total_h,
                'total_sakit': class_total_s,
                'total_izin': class_total_i,
                'total_alpa': class_total_a,
                'total_jp': class_total_jp,
                'attendance_percentage': class_attendance_percentage,
            },
            'dates': school_dates,
            'total_school_days': len(school_dates),
        }
    
    @staticmethod
    def generate_student_report(
        student: Student,
        start_date: date,
        end_date: date
    ) -> Dict:
        """
        Generate detailed attendance report for a single student.
        
        Shows dates as rows with JP columns showing individual status.
        Includes summary with totals and percentages for each status type.
        
        Args:
            student: The student to generate report for
            start_date: Start of date range
            end_date: End of date range
            
        Returns:
            Dict containing:
            - student: Student object
            - period: Dict with start_date, end_date
            - daily_records: List of daily attendance with JP details
            - summary: Totals and percentages for each status
            
        Requirements: 5.5, 5.6
        """
        classroom = student.classroom
        
        # Get all school days in the date range (excluding holidays)
        school_dates = ReportService._get_school_dates(
            classroom, start_date, end_date
        )
        
        # Get all attendance records for this student in the date range
        attendances = DailyAttendance.objects.filter(
            student=student,
            date__range=[start_date, end_date]
        ).order_by('date')
        
        # Build attendance lookup: {date: DailyAttendance}
        attendance_lookup = {att.date: att for att in attendances}
        
        # Calculate totals
        total_h = 0
        total_s = 0
        total_i = 0
        total_a = 0
        total_jp = 0
        
        # Build daily records with JP details
        daily_records = []
        
        for school_date in school_dates:
            attendance = attendance_lookup.get(school_date)
            jp_count = ScheduleService.get_jp_count_for_date(school_date)
            day_schedule = ScheduleService.get_schedule_for_date(school_date)
            
            # Build JP statuses list
            jp_details = []
            day_h = 0
            day_s = 0
            day_i = 0
            day_a = 0
            
            for jp_num in range(1, jp_count + 1):
                if attendance:
                    status = attendance.jp_statuses.get(str(jp_num), None)
                else:
                    status = None
                
                jp_details.append({
                    'jp_num': jp_num,
                    'status': status,
                    'status_display': ReportService._get_status_display(status),
                })
                
                if status == 'H':
                    day_h += 1
                    total_h += 1
                elif status == 'S':
                    day_s += 1
                    total_s += 1
                elif status == 'I':
                    day_i += 1
                    total_i += 1
                elif status == 'A':
                    day_a += 1
                    total_a += 1
                
                if status:
                    total_jp += 1
            
            daily_records.append({
                'date': school_date,
                'day_name': day_schedule.day_name if day_schedule else '',
                'jp_count': jp_count,
                'jp_details': jp_details,
                'has_record': attendance is not None,
                'day_summary': {
                    'hadir': day_h,
                    'sakit': day_s,
                    'izin': day_i,
                    'alpa': day_a,
                },
                'notes': attendance.notes if attendance else '',
            })
        
        # Calculate percentages
        attendance_percentage = round(
            (total_h / total_jp * 100), 2
        ) if total_jp > 0 else 0.0
        
        sakit_percentage = round(
            (total_s / total_jp * 100), 2
        ) if total_jp > 0 else 0.0
        
        izin_percentage = round(
            (total_i / total_jp * 100), 2
        ) if total_jp > 0 else 0.0
        
        alpa_percentage = round(
            (total_a / total_jp * 100), 2
        ) if total_jp > 0 else 0.0
        
        return {
            'student': student,
            'classroom': classroom,
            'period': {
                'start_date': start_date,
                'end_date': end_date,
            },
            'daily_records': daily_records,
            'summary': {
                'total_hadir': total_h,
                'total_sakit': total_s,
                'total_izin': total_i,
                'total_alpa': total_a,
                'total_jp': total_jp,
                'attendance_percentage': attendance_percentage,
                'sakit_percentage': sakit_percentage,
                'izin_percentage': izin_percentage,
                'alpa_percentage': alpa_percentage,
            },
            'total_school_days': len(school_dates),
            'max_jp_count': max(
                (ScheduleService.get_jp_count_for_date(d) for d in school_dates),
                default=6
            ),
        }
    
    @staticmethod
    def _get_school_dates(
        classroom: Classroom,
        start_date: date,
        end_date: date
    ) -> List[date]:
        """
        Get list of school dates (excluding holidays and non-school days).
        
        Args:
            classroom: The classroom to check holidays for
            start_date: Start of date range
            end_date: End of date range
            
        Returns:
            List of dates that are school days
        """
        school_dates = []
        current_date = start_date
        
        while current_date <= end_date:
            # Check if it's a school day according to DaySchedule
            if ScheduleService.is_school_day(current_date):
                # Check if it's not a holiday for this classroom
                if not HolidayService.is_holiday(current_date, classroom):
                    school_dates.append(current_date)
            
            current_date += timedelta(days=1)
        
        return school_dates
    
    @staticmethod
    def _get_status_display(status: Optional[str]) -> str:
        """Get display text for a status code."""
        status_map = {
            'H': 'Hadir',
            'S': 'Sakit',
            'I': 'Izin',
            'A': 'Alpa',
        }
        return status_map.get(status, '-')
    
    # ============================================
    # Legacy Report Methods (Existing)
    # ============================================
    
    @staticmethod
    def generate_attendance_report(
        start_date: date = None,
        end_date: date = None,
        classroom_id: str = None,
        status: str = None,
        page: int = 1,
        per_page: int = 50,
        **kwargs  # Accept additional kwargs for backward compatibility
    ) -> Dict:
        """Generate comprehensive attendance report with filters"""
        
        queryset = AttendanceRecord.objects.select_related(
            'student', 'student__classroom', 'student__classroom__academic_level', 'teacher'
        ).all()
        
        # Apply filters
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        if classroom_id:
            queryset = queryset.filter(student__classroom_id=classroom_id)
        if status:
            queryset = queryset.filter(status=status)
        
        # Order by date descending
        queryset = queryset.order_by('-date', 'student__classroom__academic_level', 'student__classroom__grade', 'student__name')
        
        # Pagination
        paginator = Paginator(queryset, per_page)
        records_page = paginator.get_page(page)
        
        # Generate summary statistics
        summary = ReportService._generate_report_summary(queryset)
        
        return {
            'records': records_page,
            'summary': summary,
            'filters': {
                'start_date': start_date,
                'end_date': end_date,
                'classroom_id': classroom_id,
                'status': status
            },
            'pagination': {
                'total_count': paginator.count,
                'page_count': paginator.num_pages,
                'current_page': page,
                'has_previous': records_page.has_previous(),
                'has_next': records_page.has_next(),
            }
        }
    
    @staticmethod
    def _generate_report_summary(queryset) -> Dict:
        """Generate summary statistics for a queryset"""
        total_records = queryset.count()
        
        if total_records == 0:
            return {
                'total_records': 0,
                'present': 0,
                'sick': 0,
                'permission': 0,
                'absent': 0,
                'attendance_rate': 0.0
            }
        
        present = queryset.filter(status=AttendanceStatus.HADIR).count()
        sick = queryset.filter(status=AttendanceStatus.SAKIT).count()
        permission = queryset.filter(status=AttendanceStatus.IZIN).count()
        absent = queryset.filter(status=AttendanceStatus.ALPA).count()
        
        return {
            'total_records': total_records,
            'present': present,
            'sick': sick,
            'permission': permission,
            'absent': absent,
            'attendance_rate': round((present / total_records * 100), 2)
        }
    
    @staticmethod
    def export_attendance_to_csv(
        start_date: date = None,
        end_date: date = None,
        classroom_id: str = None,
        status: str = None,
        **kwargs  # Accept additional kwargs for backward compatibility
    ) -> str:
        """Export attendance data to CSV format"""
        
        queryset = AttendanceRecord.objects.select_related(
            'student', 'student__classroom', 'student__classroom__academic_level', 'teacher'
        ).all()
        
        # Apply same filters as report
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        if classroom_id:
            queryset = queryset.filter(student__classroom_id=classroom_id)
        if status:
            queryset = queryset.filter(status=status)
        
        queryset = queryset.order_by('-date', 'student__classroom__academic_level', 'student__classroom__grade', 'student__name')
        
        # Generate CSV
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            'Tanggal', 'ID Siswa', 'Nama Siswa', 'Kelas', 'NISN', 
            'Status', 'Catatan', 'Guru', 'Waktu Input'
        ])
        
        # Data rows
        for record in queryset:
            writer.writerow([
                record.date.strftime('%Y-%m-%d'),
                record.student.student_id,
                record.student.name,
                str(record.student.classroom),
                record.student.nisn or '',
                record.get_status_display(),
                record.notes or '',
                record.teacher.get_full_name() or record.teacher.username,
                record.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return output.getvalue()
    
    @staticmethod
    def export_jp_attendance_to_csv(
        classroom: Classroom,
        start_date: date,
        end_date: date
    ) -> str:
        """
        Export JP-based attendance data to CSV format.
        
        Args:
            classroom: The classroom to export
            start_date: Start of date range
            end_date: End of date range
            
        Returns:
            CSV string content
        """
        report = ReportService.generate_class_report(classroom, start_date, end_date)
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        header = ['No', 'NIS', 'Nama Siswa']
        for school_date in report['dates']:
            header.append(school_date.strftime('%d/%m'))
        header.extend(['Total H', 'Total S', 'Total I', 'Total A', 'Total JP', 'Persentase'])
        writer.writerow(header)
        
        # Data rows
        for idx, student_data in enumerate(report['students'], 1):
            row = [
                idx,
                student_data['student'].student_id,
                student_data['student'].name,
            ]
            
            # Add daily summaries
            for daily in student_data['daily_data']:
                row.append(daily['summary'])
            
            # Add totals
            row.extend([
                student_data['total_hadir'],
                student_data['total_sakit'],
                student_data['total_izin'],
                student_data['total_alpa'],
                student_data['total_jp'],
                f"{student_data['attendance_percentage']}%"
            ])
            
            writer.writerow(row)
        
        # Summary row
        summary = report['class_summary']
        summary_row = ['', '', 'TOTAL']
        summary_row.extend(['' for _ in report['dates']])
        summary_row.extend([
            summary['total_hadir'],
            summary['total_sakit'],
            summary['total_izin'],
            summary['total_alpa'],
            summary['total_jp'],
            f"{summary['attendance_percentage']}%"
        ])
        writer.writerow(summary_row)
        
        return output.getvalue()
    
    @staticmethod
    def generate_monthly_summary(year: int, month: int) -> Dict:
        """Generate monthly attendance summary"""
        start_date = date(year, month, 1)
        
        # Calculate end date (last day of month)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        records = AttendanceRecord.objects.filter(
            date__range=[start_date, end_date]
        )
        
        # Daily statistics
        daily_stats = []
        current_date = start_date
        
        while current_date <= end_date:
            day_records = records.filter(date=current_date)
            daily_stats.append({
                'date': current_date,
                'total': day_records.count(),
                'present': day_records.filter(status=AttendanceStatus.HADIR).count(),
                'sick': day_records.filter(status=AttendanceStatus.SAKIT).count(),
                'permission': day_records.filter(status=AttendanceStatus.IZIN).count(),
                'absent': day_records.filter(status=AttendanceStatus.ALPA).count(),
            })
            current_date += timedelta(days=1)
        
        # Classroom-wise summary
        classrooms = Classroom.objects.filter(is_active=True).select_related('academic_level')
        classroom_summary = []
        
        for classroom in classrooms:
            classroom_records = records.filter(student__classroom=classroom)
            total_students = Student.objects.filter(classroom=classroom, is_active=True).count()
            
            classroom_summary.append({
                'classroom': classroom,
                'classroom_name': str(classroom),
                'total_students': total_students,
                'total_records': classroom_records.count(),
                'present': classroom_records.filter(status=AttendanceStatus.HADIR).count(),
                'sick': classroom_records.filter(status=AttendanceStatus.SAKIT).count(),
                'permission': classroom_records.filter(status=AttendanceStatus.IZIN).count(),
                'absent': classroom_records.filter(status=AttendanceStatus.ALPA).count(),
            })
        
        return {
            'period': {'year': year, 'month': month, 'start_date': start_date, 'end_date': end_date},
            'overall_summary': ReportService._generate_report_summary(records),
            'daily_stats': daily_stats,
            'classroom_summary': classroom_summary
        }
    
    @staticmethod
    def generate_student_performance_report(
        classroom_id: str = None,
        min_attendance_rate: float = None
    ) -> List[Dict]:
        """Generate student performance report based on attendance"""
        
        students = Student.objects.select_related('classroom', 'classroom__academic_level').filter(is_active=True)
        if classroom_id:
            students = students.filter(classroom_id=classroom_id)
        
        performance_data = []
        
        for student in students:
            records = AttendanceRecord.objects.filter(student=student)
            total_records = records.count()
            
            if total_records == 0:
                attendance_rate = 0.0
                present_count = 0
            else:
                present_count = records.filter(status=AttendanceStatus.HADIR).count()
                attendance_rate = round((present_count / total_records * 100), 2)
            
            # Apply minimum attendance rate filter
            if min_attendance_rate is not None and attendance_rate < min_attendance_rate:
                continue
            
            performance_data.append({
                'student': student,
                'total_records': total_records,
                'present_count': present_count,
                'attendance_rate': attendance_rate,
                'sick_count': records.filter(status=AttendanceStatus.SAKIT).count(),
                'permission_count': records.filter(status=AttendanceStatus.IZIN).count(),
                'absent_count': records.filter(status=AttendanceStatus.ALPA).count(),
            })
        
        # Sort by attendance rate (descending)
        performance_data.sort(key=lambda x: x['attendance_rate'], reverse=True)
        
        return performance_data
    
    # ============================================
    # Excel Export Methods
    # ============================================
    
    @staticmethod
    def export_jp_attendance_to_excel(
        classrooms: List[Classroom],
        start_date: date,
        end_date: date
    ) -> bytes:
        """
        Export JP-based attendance data to Excel format with advanced features.
        
        Creates separate sheets per classroom with:
        - Raw data suitable for pivot table analysis
        - SUM and COUNTIF formulas for automatic totals
        - Conditional formatting (red=Alpa, orange=Sakit, blue=Izin)
        
        Args:
            classrooms: List of classrooms to export
            start_date: Start of date range
            end_date: End of date range
            
        Returns:
            Excel file content as bytes
            
        Requirements: 6.2, 6.3, 6.4, 6.5
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import FormulaRule, CellIsRule
        
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='8B7355', end_color='8B7355', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Status colors for conditional formatting
        alpa_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')  # Light red
        sakit_fill = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')  # Light orange
        izin_fill = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')  # Light blue
        hadir_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')  # Light green
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        summary_font = Font(bold=True)
        summary_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        
        for classroom in classrooms:
            # Generate report data for this classroom
            report = ReportService.generate_class_report(classroom, start_date, end_date)
            
            # Create sheet with classroom name (max 31 chars for Excel)
            sheet_name = str(classroom)[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            # Get school dates and max JP count
            school_dates = report['dates']
            max_jp = max(
                (ScheduleService.get_jp_count_for_date(d) for d in school_dates),
                default=6
            )
            
            # Build header row
            # Columns: No, NIS, Nama, [Date1_JP1, Date1_JP2, ...], Total H, Total S, Total I, Total A, Total JP, %
            headers = ['No', 'NIS', 'Nama Siswa']
            
            # Add date/JP columns
            date_jp_columns = []
            for school_date in school_dates:
                jp_count = ScheduleService.get_jp_count_for_date(school_date)
                for jp_num in range(1, jp_count + 1):
                    col_header = f"{school_date.strftime('%d/%m')}\nJP{jp_num}"
                    headers.append(col_header)
                    date_jp_columns.append((school_date, jp_num))
            
            # Add summary columns
            headers.extend(['Total H', 'Total S', 'Total I', 'Total A', 'Total JP', 'Persentase'])
            
            # Write header row
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Set column widths
            ws.column_dimensions['A'].width = 5   # No
            ws.column_dimensions['B'].width = 12  # NIS
            ws.column_dimensions['C'].width = 25  # Nama
            
            # Date/JP columns - narrow
            for col_idx in range(4, 4 + len(date_jp_columns)):
                ws.column_dimensions[get_column_letter(col_idx)].width = 6
            
            # Summary columns
            summary_start_col = 4 + len(date_jp_columns)
            for i in range(6):
                ws.column_dimensions[get_column_letter(summary_start_col + i)].width = 10
            
            # Write data rows
            data_start_row = 2
            for row_idx, student_data in enumerate(report['students'], data_start_row):
                student = student_data['student']
                
                # Basic info
                ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
                ws.cell(row=row_idx, column=2, value=student.student_id).border = thin_border
                ws.cell(row=row_idx, column=3, value=student.name).border = thin_border
                
                # Build attendance lookup for this student
                attendance_lookup = {}
                for daily in student_data['daily_data']:
                    attendance_lookup[daily['date']] = daily
                
                # Get all attendance records for this student
                student_attendances = DailyAttendance.objects.filter(
                    student=student,
                    date__range=[start_date, end_date]
                )
                jp_status_lookup = {}
                for att in student_attendances:
                    jp_status_lookup[att.date] = att.jp_statuses
                
                # Write JP status cells
                col_idx = 4
                for school_date, jp_num in date_jp_columns:
                    jp_statuses = jp_status_lookup.get(school_date, {})
                    status = jp_statuses.get(str(jp_num), '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=status)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    col_idx += 1
                
                # Summary columns with formulas
                summary_col = col_idx
                data_range_start = get_column_letter(4)
                data_range_end = get_column_letter(col_idx - 1)
                data_range = f"{data_range_start}{row_idx}:{data_range_end}{row_idx}"
                
                # Total H (COUNTIF formula)
                cell = ws.cell(row=row_idx, column=summary_col, 
                              value=f'=COUNTIF({data_range},"H")')
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                
                # Total S
                cell = ws.cell(row=row_idx, column=summary_col + 1,
                              value=f'=COUNTIF({data_range},"S")')
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                
                # Total I
                cell = ws.cell(row=row_idx, column=summary_col + 2,
                              value=f'=COUNTIF({data_range},"I")')
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                
                # Total A
                cell = ws.cell(row=row_idx, column=summary_col + 3,
                              value=f'=COUNTIF({data_range},"A")')
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                
                # Total JP (count non-empty cells)
                cell = ws.cell(row=row_idx, column=summary_col + 4,
                              value=f'=COUNTA({data_range})')
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                
                # Percentage
                h_col = get_column_letter(summary_col)
                jp_col = get_column_letter(summary_col + 4)
                cell = ws.cell(row=row_idx, column=summary_col + 5,
                              value=f'=IF({jp_col}{row_idx}>0,ROUND({h_col}{row_idx}/{jp_col}{row_idx}*100,2),0)')
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.number_format = '0.00"%"'
            
            # Add summary row
            last_data_row = data_start_row + len(report['students']) - 1
            summary_row = last_data_row + 2
            
            ws.cell(row=summary_row, column=1, value='').border = thin_border
            ws.cell(row=summary_row, column=2, value='').border = thin_border
            cell = ws.cell(row=summary_row, column=3, value='TOTAL')
            cell.font = summary_font
            cell.fill = summary_fill
            cell.border = thin_border
            
            # Empty cells for date columns
            for col_idx in range(4, 4 + len(date_jp_columns)):
                cell = ws.cell(row=summary_row, column=col_idx, value='')
                cell.fill = summary_fill
                cell.border = thin_border
            
            # Summary totals with SUM formulas
            summary_col = 4 + len(date_jp_columns)
            for i in range(6):
                col_letter = get_column_letter(summary_col + i)
                cell = ws.cell(row=summary_row, column=summary_col + i,
                              value=f'=SUM({col_letter}{data_start_row}:{col_letter}{last_data_row})')
                cell.font = summary_font
                cell.fill = summary_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            
            # Fix percentage formula for summary row
            h_col = get_column_letter(summary_col)
            jp_col = get_column_letter(summary_col + 4)
            cell = ws.cell(row=summary_row, column=summary_col + 5,
                          value=f'=IF({jp_col}{summary_row}>0,ROUND({h_col}{summary_row}/{jp_col}{summary_row}*100,2),0)')
            cell.font = summary_font
            cell.fill = summary_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '0.00"%"'
            
            # Apply conditional formatting to status cells
            if date_jp_columns:
                status_range_start = f"D2"
                status_range_end = f"{get_column_letter(3 + len(date_jp_columns))}{last_data_row}"
                status_range = f"{status_range_start}:{status_range_end}"
                
                # Alpa - Red
                ws.conditional_formatting.add(
                    status_range,
                    CellIsRule(operator='equal', formula=['"A"'], fill=alpa_fill)
                )
                
                # Sakit - Orange
                ws.conditional_formatting.add(
                    status_range,
                    CellIsRule(operator='equal', formula=['"S"'], fill=sakit_fill)
                )
                
                # Izin - Blue
                ws.conditional_formatting.add(
                    status_range,
                    CellIsRule(operator='equal', formula=['"I"'], fill=izin_fill)
                )
                
                # Hadir - Green
                ws.conditional_formatting.add(
                    status_range,
                    CellIsRule(operator='equal', formula=['"H"'], fill=hadir_fill)
                )
            
            # Freeze panes (freeze first row and first 3 columns)
            ws.freeze_panes = 'D2'
            
            # Add report info at the top (insert rows)
            ws.insert_rows(1, 3)
            ws.cell(row=1, column=1, value=f'Laporan Absensi JP - {classroom}')
            ws.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws.cell(row=2, column=1, value=f'Periode: {start_date.strftime("%d/%m/%Y")} - {end_date.strftime("%d/%m/%Y")}')
            ws.cell(row=3, column=1, value=f'Total Siswa: {report["class_summary"]["total_students"]} | Total Hari Sekolah: {report["total_school_days"]}')
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
